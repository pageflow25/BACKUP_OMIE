from django.contrib import admin
from django.utils.html import format_html, mark_safe
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from .models import (
    CategoriaCadastro, ClientesCadastro, ContaCorrenteCadastro,
    ContaPagarCadastro, ContaPagarDistribuicao, ContaReceberCadastro,
    ContaReceberDistribuicao, DocumentosXml, FamiliasCadastro,
    LocaisCadastro, MovimentosFinanceiros, NfCadastro,
    NfCadastroItens, NfseEncontrada, PedidoVendaItens,
    PedidoVendaProduto, ProjetosCadastro, VendedoresCadastro
)


# Função auxiliar para exportar para Excel
def export_to_excel(modeladmin, request, queryset):
    """Exporta os registros selecionados para Excel com otimização de queries"""
    import re
    from datetime import datetime, date
    
    model = queryset.model
    model_name = model._meta.verbose_name_plural or model._meta.model_name
    
    # Otimizar queryset com select_related para ForeignKeys
    related_fields = []
    for field in model._meta.fields:
        if field.is_relation and field.related_model:
            related_fields.append(field.name)
    
    if related_fields:
        queryset = queryset.select_related(*related_fields)
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = model_name[:31]
    
    # Obter campos do list_display ou todos os campos do model
    if hasattr(modeladmin, 'list_display') and modeladmin.list_display:
        field_names = []
        headers = []
        for field in modeladmin.list_display:
            if field == '__str__':
                continue
            if hasattr(model, field):
                field_names.append(field)
                try:
                    headers.append(model._meta.get_field(field).verbose_name.title())
                except:
                    headers.append(field.replace('_', ' ').title())
            elif hasattr(modeladmin, field):
                field_names.append(field)
                method = getattr(modeladmin, field)
                if hasattr(method, 'short_description'):
                    headers.append(method.short_description)
                else:
                    headers.append(field.replace('_', ' ').title())
    else:
        field_names = [f.name for f in model._meta.fields]
        headers = [f.verbose_name.title() for f in model._meta.fields]
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Escrever cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Rastrear larguras máximas das colunas
    col_widths = [len(str(h)) for h in headers]
    
    # Escrever dados usando iterator() para economia de memória
    row_num = 2
    for obj in queryset.iterator():
        for col_num, field in enumerate(field_names, 1):
            try:
                if hasattr(model, field) and hasattr(getattr(model, field, None), 'field'):
                    value = getattr(obj, field, '')
                elif hasattr(modeladmin, field):
                    method = getattr(modeladmin, field)
                    value = method(obj)
                    if isinstance(value, str) and '<' in value:
                        value = re.sub('<[^<]+?>', '', value)
                else:
                    value = getattr(obj, field, '')
                
                # Converter valores
                if value is None:
                    value = ''
                elif isinstance(value, (datetime, date)):
                    value = value.strftime('%d/%m/%Y') if isinstance(value, date) else value.strftime('%d/%m/%Y %H:%M')
                elif hasattr(value, '__str__') and not isinstance(value, (str, int, float, bool)):
                    value = str(value)
                
                ws.cell(row=row_num, column=col_num, value=value)
                
                # Atualizar largura máxima (apenas primeiras 500 linhas para performance)
                if row_num <= 500:
                    col_widths[col_num - 1] = max(col_widths[col_num - 1], len(str(value)[:50]))
            except Exception:
                ws.cell(row=row_num, column=col_num, value='')
        row_num += 1
    
    # Ajustar largura das colunas
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = min(width + 2, 50)
    
    # Criar resposta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{model_name}.xlsx"'
    wb.save(response)
    return response

export_to_excel.short_description = "📊 Exportar selecionados para Excel"


# Configuração para CategoriaCadastro
@admin.register(CategoriaCadastro)
class CategoriaCadastroAdmin(admin.ModelAdmin):
    list_display = ['id', 'codigo_formatado', 'descricao', 'tipo_categoria', 'natureza', 'status_conta']
    list_filter = [
        'tipo_categoria', 
        'natureza', 
        'conta_inativa', 
        'transferencia',
        'totalizadora',
        'nao_exibir',
        'definida_pelo_usuario',
        ('conta_receita', admin.EmptyFieldListFilter),
        ('conta_despesa', admin.EmptyFieldListFilter),
        ('codigo_dre', admin.EmptyFieldListFilter)
    ]
    search_fields = ['codigo', 'descricao', 'descricao_padrao']
    list_per_page = 25
    ordering = ['codigo']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    fieldsets = (
        ('📁 Informações Básicas', {
            'fields': (
                ('codigo', 'categoria_superior'),
                ('descricao', 'descricao_padrao'),
                ('tipo_categoria', 'natureza'),
            ),
            'classes': ('wide',),
            'description': 'Dados fundamentais da categoria contábil'
        }),
        ('⚙️ Configurações Oper.', {
            'fields': (
                ('conta_inativa', 'transferencia'),
                ('totalizadora', 'nao_exibir'),
                ('definida_pelo_usuario',),
            ),
            'classes': ('wide',)
        }),
        ('🏦 Contas Contábeis', {
            'fields': (
                ('conta_receita', 'conta_despesa'),
                ('id_conta_contabil', 'tag_conta_contabil'),
                ('codigo_dre',),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('📈 DRE - Demonstração', {
            'fields': (
                ('dadosdre_codigodre', 'dadosdre_descricaodre'),
                ('dadosdre_niveldre', 'dadosdre_sinaldre'),
                ('dadosdre_totalizadre', 'dadosdre_naoexibirdre'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('🕰️ Sistema', {
            'fields': (
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )
    
    @admin.display(description='Código')
    def codigo_formatado(self, obj):
        if obj.codigo:
            try:
                # Verificar se é um número
                codigo_num = float(obj.codigo)
                # Formatar como inteiro se for um número inteiro, senão como float
                if codigo_num == int(codigo_num):
                    return str(int(codigo_num))
                return f'{codigo_num:.2f}'
            except (ValueError, TypeError):
                # Se não for um número, retornar como string
                return str(obj.codigo)
        return '-'
    
    @admin.display(description='Status')
    def status_conta(self, obj):
        if obj.conta_inativa == 'S':
            return mark_safe('<span style="color: red;">● Inativa</span>')
        return mark_safe('<span style="color: green;">● Ativa</span>')


# Configuração para ClientesCadastro
@admin.register(ClientesCadastro)
class ClientesCadastroAdmin(admin.ModelAdmin):
    list_display = ['codigo_cliente_omie', 'razao_social', 'nome_fantasia', 'cnpj_cpf', 'cidade', 'estado', 'status_cliente']
    list_filter = [
        'estado', 
        'inativo', 
        'pessoa_fisica', 
        'contribuinte', 
        'bloquear_faturamento',
        'tags',
        ('email', admin.EmptyFieldListFilter)
    ]
    search_fields = ['razao_social', 'nome_fantasia', 'cnpj_cpf', 'codigo_cliente_integracao', 'email']
    list_per_page = 25
    ordering = ['razao_social']
    
    @admin.display(description='Status')
    def status_cliente(self, obj):
        if obj.inativo == 'S':
            return mark_safe('<span style="color: red;">● Inativo</span>')
        if obj.bloquear_faturamento == 'S':
            return mark_safe('<span style="color: orange;">● Bloqueado</span>')
        return mark_safe('<span style="color: green;">● Ativo</span>')
    
    fieldsets = (
        ('👤 Identificação', {
            'fields': (
                ('codigo_cliente_omie', 'codigo_cliente_integracao'),
                ('razao_social',),
                ('nome_fantasia', 'contato'),
            ),
            'classes': ('wide',),
            'description': 'Dados básicos de identificação do cliente'
        }),
        ('📋 Documento e Inscrições', {
            'fields': (
                ('cnpj_cpf', 'pessoa_fisica'),
                ('inscricao_estadual', 'inscricao_municipal'),
                ('contribuinte', 'optante_simples_nacional'),
                ('cnae', 'produtor_rural'),
            ),
            'classes': ('wide',)
        }),
        ('📍 Endereço Principal', {
            'fields': (
                ('endereco', 'endereco_numero'),
                ('complemento', 'bairro'),
                ('cidade', 'estado'),
                ('cep', 'codigo_pais'),
                ('exterior',),
            ),
            'classes': ('wide',)
        }),
        ('📦 Endereço de Entrega', {
            'fields': (
                ('enderecoentrega_entendereco', 'enderecoentrega_entnumero'),
                ('enderecoentrega_entcomplemento', 'enderecoentrega_entbairro'),
                ('enderecoentrega_entcidade', 'enderecoentrega_entestado'),
                ('enderecoentrega_entcep',),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('📞 Contato', {
            'fields': (
                ('email', 'homepage'),
                ('telefone1_ddd', 'telefone1_numero'),
                ('telefone2_ddd', 'telefone2_numero'),
                ('fax_ddd', 'fax_numero'),
            ),
            'classes': ('wide',)
        }),
        ('🏦 Dados Bancários', {
            'fields': (
                ('dadosbancarios_codigo_banco', 'dadosbancarios_agencia'),
                ('dadosbancarios_conta_corrente',),
                ('dadosbancarios_nome_titular', 'dadosbancarios_doc_titular'),
                ('dadosbancarios_cchavepix',),
                ('dadosbancarios_transf_padrao',),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('💼 Configurações Comerciais', {
            'fields': (
                ('recomendacoes_codigo_vendedor', 'recomendacoes_tipo_assinante'),
                ('recomendacoes_numero_parcelas',),
                ('recomendacoes_gerar_boletos',),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('🅰️ Status e Controle', {
            'fields': (
                ('inativo',),
                ('bloquear_faturamento', 'bloquear_exclusao'),
                ('enviar_anexos',),
            ),
            'classes': ('wide',)
        }),
        ('🏷️ Tags', {
            'fields': (
                ('tags_0_tag', 'tags_1_tag'),
                ('tags_2_tag',),
                ('tags', 'tags_json'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )


# Configuração para ContaCorrenteCadastro
@admin.register(ContaCorrenteCadastro)
class ContaCorrenteCadastroAdmin(admin.ModelAdmin):
    list_display = ['ncodcc', 'descricao', 'codigo_banco_formatado', 'codigo_agencia', 'numero_conta_corrente', 'tipo', 'status_conta']
    list_filter = [
        'codigo_banco', 
        'inativo', 
        'tipo', 
        'bloqueado',
        'tipo_conta_corrente',
        'modalidade',
        'nao_fluxo',
        'nao_resumo',
        'importado_api',
        ('codigo_agencia', admin.EmptyFieldListFilter),
        ('email', admin.EmptyFieldListFilter)
    ]
    search_fields = ['descricao', 'numero_conta_corrente', 'codigo_agencia', 'ncodcc']
    list_per_page = 25
    ordering = ['descricao']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    fieldsets = (
        ('Identificação', {
            'fields': (
                ('ncodcc', 'ccodccint'),
                ('descricao', 'tipo'),
                ('tipo_conta_corrente',),
            ),
            'classes': ('wide',)
        }),
        ('Dados Bancários', {
            'fields': (
                ('codigo_banco', 'codigo_agencia'),
                ('numero_conta_corrente', 'modalidade'),
                ('cestabelecimento', 'ctipocartao'),
            ),
            'classes': ('wide',)
        }),
        ('Contato e Endereço', {
            'fields': (
                ('endereco', 'numero'),
                ('complemento', 'bairro'),
                ('cidade', 'estado'),
                ('cep', 'codigo_pais'),
                ('ddd', 'telefone'),
                ('email', 'nome_gerente'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('Configurações Financeiras', {
            'fields': (
                ('saldo_inicial', 'valor_limite'),
                ('saldo_data', 'per_juros'),
                ('per_multa', 'dias_rcomp'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('Status e Controles', {
            'fields': (
                ('inativo', 'bloqueado'),
                ('nao_fluxo', 'nao_resumo'),
                ('importado_api',),
            ),
            'classes': ('wide',)
        }),
        ('Sistema', {
            'fields': (
                ('data_inc', 'hora_inc', 'user_inc'),
                ('data_alt', 'hora_alt', 'user_alt'),
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )
    
    @admin.display(description='Banco')
    def codigo_banco_formatado(self, obj):
        if obj.codigo_banco:
            try:
                # Verificar se é um número
                banco_num = float(obj.codigo_banco)
                # Formatar como inteiro se for um número inteiro, senão como float
                if banco_num == int(banco_num):
                    return str(int(banco_num))
                return f'{banco_num:.2f}'
            except (ValueError, TypeError):
                # Se não for um número, retornar como string
                return str(obj.codigo_banco)
        return '-'
    
    @admin.display(description='Status')
    def status_conta(self, obj):
        if obj.inativo == 'S':
            return mark_safe('<span style="color: red;">● Inativa</span>')
        if obj.bloqueado == 'S':
            return mark_safe('<span style="color: orange;">● Bloqueada</span>')
        return mark_safe('<span style="color: green;">● Ativa</span>')


# Configuração para ContaPagarCadastro  
@admin.register(ContaPagarCadastro)
class ContaPagarCadastroAdmin(admin.ModelAdmin):
    list_display = ['codigo_lancamento_omie', 'nome_cliente', 'nome_vendedor', 'nome_projeto', 'numero_documento', 'valor_formatado', 'data_vencimento', 'status_visual']
    list_filter = [
        'status_titulo', 
        ('data_emissao', admin.DateFieldListFilter), 
        ('data_vencimento', admin.DateFieldListFilter), 
        'retem_ir', 
        'retem_iss', 
        'bloqueado',
        'codigo_categoria',
        ('numero_documento_fiscal', admin.EmptyFieldListFilter)
    ]
    search_fields = ['numero_documento', 'codigo_lancamento_integracao', 'cliente__razao_social', 'vendedor__nome']
    list_per_page = 25
    ordering = ['-data_vencimento']
    autocomplete_fields = ['cliente', 'vendedor', 'projeto']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    fieldsets = (
        ('🆔 Identificação do Título', {
            'fields': (
                ('codigo_lancamento_omie', 'codigo_lancamento_integracao'),
                ('numero_documento', 'numero_documento_fiscal'),
                ('chave_nfe', 'codigo_barras_ficha_compensacao'),
            ),
            'classes': ('wide',),
            'description': 'Códigos identificadores únicos do lançamento'
        }),
        ('🔗 Relacionamentos', {
            'fields': (
                ('cliente', 'vendedor'),
                ('projeto', 'id_conta_corrente'),
            ),
            'classes': ('wide',)
        }),
        ('💰 Informações Financeiras', {
            'fields': (
                ('valor_documento', 'numero_parcela'),
                ('codigo_categoria', 'operacao'),
                ('numero_pedido',),
            ),
            'classes': ('wide',)
        }),
        ('📅 Cronograma e Prazos', {
            'fields': (
                ('data_emissao', 'data_entrada'),
                ('data_previsao', 'data_vencimento'),
            ),
            'classes': ('wide',)
        }),
        ('⚠️ Status e Controles', {
            'fields': (
                ('status_titulo', 'bloqueado'),
                ('baixa_bloqueada', 'bloquear_exclusao'),
                ('id_origem',),
            ),
            'classes': ('wide',)
        }),
        ('📊 Retenções de Impostos', {
            'fields': (
                ('retem_ir', 'retem_iss'),
                ('retem_inss', 'retem_cofins'),
                ('retem_csll', 'retem_pis'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('Sistema', {
            'fields': (
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )
    
    @admin.display(description='Cliente/Fornecedor')
    def nome_cliente(self, obj):
        if obj.cliente:
            return obj.cliente.razao_social or obj.cliente.nome_fantasia
        return f'Cód: {obj.codigo_cliente_fornecedor}' if obj.codigo_cliente_fornecedor else '-'
    
    @admin.display(description='Vendedor')
    def nome_vendedor(self, obj):
        if obj.vendedor:
            return obj.vendedor.nome
        return '-'
    
    @admin.display(description='Projeto')
    def nome_projeto(self, obj):
        if obj.projeto:
            return obj.projeto.nome
        return '-'
    
    @admin.display(description='Valor')
    def valor_formatado(self, obj):
        if obj.valor_documento:
            return f'R$ {obj.valor_documento:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return '-'
    
    @admin.display(description='Status')
    def status_visual(self, obj):
        status_colors = {
            'LIQUIDADO': 'green',
            'ABERTO': 'orange',
            'ATRASADO': 'red',
            'CANCELADO': 'gray',
        }
        color = status_colors.get(obj.status_titulo, 'blue')
        return format_html('<span style="color: {};">{}</span>', color, obj.status_titulo or '-')


# Configuração para ContaPagarDistribuicao
@admin.register(ContaPagarDistribuicao)
class ContaPagarDistribuicaoAdmin(admin.ModelAdmin):
    list_display = ['id', 'parent_id', 'item_index', 'ccoddep', 'cdesdep', 'nvaldep']
    list_filter = ['ccoddep']
    search_fields = ['cdesdep']
    list_per_page = 20


# Configuração para ContaReceberCadastro
@admin.register(ContaReceberCadastro)
class ContaReceberCadastroAdmin(admin.ModelAdmin):
    list_display = ['codigo_lancamento_omie', 'nome_cliente', 'nome_vendedor', 'nome_projeto', 'numero_documento', 'valor_formatado', 'data_vencimento', 'status_visual']
    list_filter = [
        'status_titulo', 
        ('data_emissao', admin.DateFieldListFilter), 
        ('data_vencimento', admin.DateFieldListFilter), 
        'vendedor_rel', 
        'bloqueado',
        'codigo_categoria',
        ('numero_documento_fiscal', admin.EmptyFieldListFilter)
    ]
    search_fields = ['numero_documento', 'codigo_lancamento_integracao', 'cliente__razao_social', 'vendedor_rel__nome']
    list_per_page = 25
    ordering = ['-data_vencimento']
    autocomplete_fields = ['cliente', 'vendedor_rel', 'projeto_rel']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    fieldsets = (
        ('🆔 Identificação da Conta', {
            'fields': (
                ('codigo_lancamento_omie', 'codigo_lancamento_integracao'),
                ('numero_documento', 'numero_documento_fiscal'),
                ('chave_nfe', 'codigo_barras_ficha_compensacao'),
            ),
            'classes': ('wide',),
            'description': 'Identificadores únicos e documentos vinculados'
        }),
        ('🔗 Vínculos Comerciais', {
            'fields': (
                ('cliente', 'vendedor_rel'),
                ('projeto_rel', 'id_conta_corrente'),
            ),
            'classes': ('wide',)
        }),
        ('💵 Valores e Impostos', {
            'fields': (
                ('valor_documento', 'valor_iss'),
                ('numero_parcela', 'numero_pedido'),
                ('codigo_categoria', 'operacao'),
                ('ncodos', 'tipo_agrupamento'),
            ),
            'classes': ('wide',)
        }),
        ('📆 Prazos de Pagamento', {
            'fields': (
                ('data_emissao', 'data_registro'),
                ('data_previsao', 'data_vencimento'),
            ),
            'classes': ('wide',)
        }),
        ('⛔ Status e Restrições', {
            'fields': (
                ('status_titulo', 'bloqueado'),
                ('bloquear_baixa', 'bloquear_exclusao'),
                ('id_origem',),
            ),
            'classes': ('wide',)
        }),
        ('🏦 Dados do Boleto', {
            'fields': (
                ('boleto_cgerado', 'boleto_cnumboleto'),
                ('boleto_cnumbancario', 'boleto_ddtembol'),
                ('boleto_nperjuros', 'boleto_npermulta'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('Retenções', {
            'fields': (
                ('retem_ir', 'retem_iss'),
                ('retem_inss', 'retem_cofins'),
                ('retem_csll', 'retem_pis'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('Sistema', {
            'fields': (
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )
    
    @admin.display(description='Cliente')
    def nome_cliente(self, obj):
        if obj.cliente:
            return obj.cliente.razao_social or obj.cliente.nome_fantasia
        return f'Cód: {obj.codigo_cliente_fornecedor}' if obj.codigo_cliente_fornecedor else '-'
    
    @admin.display(description='Vendedor')
    def nome_vendedor(self, obj):
        if obj.vendedor_rel:
            return obj.vendedor_rel.nome
        return '-'
    
    @admin.display(description='Projeto')
    def nome_projeto(self, obj):
        if obj.projeto_rel:
            return obj.projeto_rel.nome
        return '-'
    
    @admin.display(description='Valor')
    def valor_formatado(self, obj):
        if obj.valor_documento:
            return f'R$ {obj.valor_documento:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return '-'
    
    @admin.display(description='Status')
    def status_visual(self, obj):
        status_colors = {
            'LIQUIDADO': 'green',
            'ABERTO': 'orange',
            'ATRASADO': 'red',
            'CANCELADO': 'gray',
        }
        color = status_colors.get(obj.status_titulo, 'blue')
        return format_html('<span style="color: {};">{}</span>', color, obj.status_titulo or '-')


# Configuração para ContaReceberDistribuicao
@admin.register(ContaReceberDistribuicao)
class ContaReceberDistribuicaoAdmin(admin.ModelAdmin):
    list_display = ['id', 'parent_id', 'item_index', 'ccoddep', 'cdesdep', 'nvaldep']
    list_filter = ['ccoddep']
    search_fields = ['cdesdep']
    list_per_page = 20


# Configuração para DocumentosXml
@admin.register(DocumentosXml)
class DocumentosXmlAdmin(admin.ModelAdmin):
    list_display = ['nidnf', 'nnumero', 'cserie', 'nvalor', 'demissao', 'cstatus']
    list_filter = ['cstatus', 'demissao', 'cserie']
    search_fields = ['nnumero', 'nchave']
    list_per_page = 20
    ordering = ['-demissao']


# Configuração para FamiliasCadastro
@admin.register(FamiliasCadastro)
class FamiliasCadastroAdmin(admin.ModelAdmin):
    list_display = ['codigo', 'codfamilia_formatada', 'nomefamilia', 'codint', 'inativo']
    list_filter = ['inativo']
    search_fields = ['nomefamilia', 'codint']
    list_per_page = 20
    ordering = ['nomefamilia']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    fieldsets = (
        ('📁 Dados da Família', {
            'fields': (
                ('codigo', 'codfamilia'),
                ('nomefamilia', 'codint'),
                ('inativo',),
            ),
            'classes': ('wide',),
            'description': 'Classificação e agrupamento de produtos'
        }),
        ('🕰️ Sistema', {
            'fields': (
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )
    
    @admin.display(description='Cód. Família')
    def codfamilia_formatada(self, obj):
        if obj.codfamilia:
            try:
                # Verificar se é um número
                codigo_num = float(obj.codfamilia)
                # Formatar como inteiro se for um número inteiro, senão como float
                if codigo_num == int(codigo_num):
                    return str(int(codigo_num))
                return f'{codigo_num:.2f}'
            except (ValueError, TypeError):
                # Se não for um número, retornar como string
                return str(obj.codfamilia)
        return '-'


# Configuração para LocaisCadastro
@admin.register(LocaisCadastro)
class LocaisCadastroAdmin(admin.ModelAdmin):
    list_display = ['codigo_local_estoque', 'codigo', 'descricao', 'tipo_formatado', 'padrao', 'inativo']
    list_filter = [
        'tipo', 
        'padrao', 
        'inativo', 
        'dispvenda', 
        'dispordemproducao',
        'dispconsumoop',
        'dispremessa'
    ]
    search_fields = ['descricao', 'codigo']
    list_per_page = 20
    ordering = ['descricao']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    fieldsets = (
        ('📍 Local de Estoque', {
            'fields': (
                ('codigo_local_estoque', 'codigo'),
                ('descricao', 'tipo'),
                ('padrao', 'inativo'),
            ),
            'classes': ('wide',),
            'description': 'Identificação e tipo do local'
        }),
        ('📦 Disponibilidade', {
            'fields': (
                ('dispvenda', 'dispordemproducao'),
                ('dispconsumoop', 'dispremessa'),
            ),
            'classes': ('wide',)
        }),
        ('🕰️ Controle de Sistema', {
            'fields': (
                ('dinc', 'hinc', 'uinc'),
                ('dalt', 'halt', 'ualt'),
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )
    
    @admin.display(description='Tipo')
    def tipo_formatado(self, obj):
        if obj.tipo:
            try:
                # Verificar se é um número
                tipo_num = float(obj.tipo)
                # Formatar como inteiro se for um número inteiro, senão como float
                if tipo_num == int(tipo_num):
                    return str(int(tipo_num))
                return f'{tipo_num:.2f}'
            except (ValueError, TypeError):
                # Se não for um número, retornar como string
                return str(obj.tipo)
        return '-'


# Configuração para MovimentosFinanceiros
@admin.register(MovimentosFinanceiros)
class MovimentosFinanceirosAdmin(admin.ModelAdmin):
    list_display = ['id', 'detalhes_cnumtitulo', 'nome_cliente', 'nome_conta_corrente', 'nome_vendedor', 'nome_categoria', 'valor_formatado', 'detalhes_ddtvenc', 'detalhes_ddtpagamento', 'status_visual']
    list_filter = ['detalhes_cstatus', 'detalhes_corigem', 'detalhes_cnatureza', 'detalhes_ccodcateg']
    search_fields = ['detalhes_cnumtitulo', 'cliente__razao_social', 'conta_corrente__descricao']
    list_per_page = 25
    ordering = ['-detalhes_ddtvenc']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    autocomplete_fields = ['cliente', 'conta_corrente', 'vendedor', 'projeto']
    list_select_related = ['cliente', 'conta_corrente', 'vendedor', 'projeto']
    actions = [export_to_excel]
    
    fieldsets = (
        ('Informações Principais', {
            'fields': (
                ('detalhes_cnumtitulo', 'detalhes_ncodtitulo'),
                ('cliente', 'conta_corrente'),
                ('vendedor', 'projeto'),
            ),
            'classes': ('wide',)
        }),
        ('📊 Status e Origem', {
            'fields': (
                ('detalhes_cstatus', 'detalhes_corigem'),
                ('detalhes_cnatureza', 'detalhes_ctipo'),
                ('detalhes_cgrupo', 'detalhes_coperacao'),
            ),
            'classes': ('wide',)
        }),
        ('💵 Valores Principais', {
            'fields': (
                ('detalhes_nvalortitulo', 'detalhes_cnumparcela'),
                ('detalhes_ccodcateg', 'detalhes_ncodos'),
                ('detalhes_ncodnf', 'detalhes_ncodtitrepet'),
            ),
            'classes': ('wide',)
        }),
        ('📅 Cronograma Financeiro', {
            'fields': (
                ('detalhes_ddtemissao', 'detalhes_ddtregistro'),
                ('detalhes_ddtprevisao', 'detalhes_ddtvenc'),
                ('detalhes_ddtpagamento', 'detalhes_ddtcredito'),
                ('detalhes_ddtconcilia',),
            ),
            'classes': ('wide',)
        }),
        ('💰 Valores Detalhados', {
            'fields': (
                ('resumo_nvalliquido', 'resumo_nvalpago'),
                ('resumo_nvalaberto', 'resumo_ndesconto'),
                ('resumo_njuros', 'resumo_nmulta'),
                ('detalhes_nmulta', 'detalhes_njuros'),
                ('detalhes_ndesconto',),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('📄 Documentos e Códigos', {
            'fields': (
                ('detalhes_cnumdocfiscal', 'detalhes_ccodinttitulo'),
                ('detalhes_ccodigobarras', 'detalhes_cnumboleto'),
                ('detalhes_cchavenfe', 'detalhes_cretiss'),
                ('detalhes_nvaloriss', 'detalhes_ncodcomprador'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('Controle de Sistema', {
            'fields': (
                ('detalhes_ccpfcnpjcliente', 'detalhes_chrconcilia'),
                ('resumo_cliquidado',),
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('Movimentação Conta Corrente', {
            'fields': (
                ('detalhes_ncodbaixa', 'detalhes_ncodmovcc'),
                ('detalhes_ncodmovccrepet', 'detalhes_nvalormovcc'),
            ),
            'classes': ('wide', 'collapse')
        })
    )
    
    @admin.display(description='Cliente')
    def nome_cliente(self, obj):
        if obj.cliente:
            return obj.cliente.razao_social or obj.cliente.nome_fantasia
        return '-'
    
    @admin.display(description='Conta Corrente')
    def nome_conta_corrente(self, obj):
        if obj.conta_corrente:
            return obj.conta_corrente.descricao
        return '-'
    
    @admin.display(description='Vendedor')
    def nome_vendedor(self, obj):
        if obj.vendedor:
            return obj.vendedor.nome
        return '-'
    
    @admin.display(description='Categoria')
    def nome_categoria(self, obj):
        if obj.detalhes_ccodcateg:
            try:
                # Tentar buscar por codigo_dre primeiro, depois por codigo
                categoria = CategoriaCadastro.objects.filter(codigo_dre=obj.detalhes_ccodcateg).first()
                if not categoria:
                    categoria = CategoriaCadastro.objects.filter(codigo=obj.detalhes_ccodcateg).first()
                if categoria:
                    return categoria.descricao or obj.detalhes_ccodcateg
            except Exception:
                pass
            return obj.detalhes_ccodcateg
        return '-'
    
    @admin.display(description='Valor')
    def valor_formatado(self, obj):
        if obj.detalhes_nvalortitulo:
            return f'R$ {obj.detalhes_nvalortitulo:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return '-'
    
    @admin.display(description='Status')
    def status_visual(self, obj):
        status_colors = {
            'LIQUIDADO': 'green',
            'ABERTO': 'orange', 
            'ATRASADO': 'red',
            'CANCELADO': 'gray',
        }
        color = status_colors.get(obj.detalhes_cstatus, 'blue')
        return format_html('<span style="color: {};">{}</span>', color, obj.detalhes_cstatus or '-')


# Configuração para NfCadastro
@admin.register(NfCadastro)
class NfCadastroAdmin(admin.ModelAdmin):
    list_display = ['nidnf', 'ide_nnf', 'destinatario_nome', 'total_icmstot_vnf', 'ide_diemi']
    list_filter = [
        ('ide_diemi', admin.DateFieldListFilter),
        ('ide_demi', admin.DateFieldListFilter),
        ('ide_dsaient', admin.DateFieldListFilter),
        'ide_mod', 
        'ide_tpnf',
        ('ide_dcan', admin.DateFieldListFilter),
        ('ide_dinut', admin.DateFieldListFilter),
        ('destinatario_cnpjcpf', admin.EmptyFieldListFilter)
    ]
    search_fields = ['ide_nnf', 'destinatario_nome', 'destinatario_cnpjcpf']
    list_per_page = 20
    ordering = ['-ide_diemi']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    fieldsets = (
        ('🏷️ Identificação da NF-e', {
            'fields': (
                ('nidnf', 'ccodintnf'),
                ('ide_nnf', 'ide_srie'),
                ('ide_mod', 'ide_tpnf'),
            ),
            'classes': ('wide',),
            'description': 'Números e identificações da nota fiscal'
        }),
        ('📅 Datas e Horários', {
            'fields': (
                ('ide_diemi', 'ide_demi'),
                ('ide_hemi', 'ide_dsaient'),
                ('ide_hsaient', 'ide_dreg'),
                ('ide_dcan', 'ide_dinut'),
            ),
            'classes': ('wide',)
        }),
        ('👤 Destinatário', {
            'fields': (
                ('destinatario_nome', 'destinatario_cnpjcpf'),
                ('nfdestint_crazao', 'nfdestint_cnpj_cpf'),
                ('nfdestint_ncodcli', 'nfdestint_ccodcliint'),
            ),
            'classes': ('wide',)
        }),
        ('💰 Totais da NF-e', {
            'fields': (
                ('total_icmstot_vnf', 'total_icmstot_vprod'),
                ('total_icmstot_vicms', 'total_icmstot_vst'),
                ('total_icmstot_vdesc', 'total_icmstot_vfrete'),
                ('total_icmstot_vseg', 'total_icmstot_voutro'),
            ),
            'classes': ('wide', 'collapse')
        }),
        ('🕰️ Sistema', {
            'fields': (
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )


# Configuração para NfCadastroItens (Inline para NfCadastro)
class NfCadastroItensInline(admin.TabularInline):
    model = NfCadastroItens
    extra = 0
    readonly_fields = ['sync_created_at', 'sync_updated_at']


# Configuração para NfCadastroItens
@admin.register(NfCadastroItens)
class NfCadastroItensAdmin(admin.ModelAdmin):
    list_display = ['id', 'parent_id', 'item_index', 'prod_xprod', 'prod_vprod', 'prod_qcom']
    list_filter = ['prod_cfop', 'prod_ncm']
    search_fields = ['prod_xprod', 'prod_cprod']
    list_per_page = 20
    readonly_fields = ['sync_created_at', 'sync_updated_at']


# Configuração para NfseEncontrada
@admin.register(NfseEncontrada)
class NfseEncontradaAdmin(admin.ModelAdmin):
    list_display = ['id', 'cabecalho_ncodnf', 'cabecalho_crazaodestinatario', 'cabecalho_nvalornfse', 'emissao_cdataemissao']
    list_filter = [
        'cabecalho_cstatusnfse', 
        ('emissao_cdataemissao', admin.DateFieldListFilter),
        ('cabecalho_crazaodestinatario', admin.EmptyFieldListFilter)
    ]
    search_fields = ['cabecalho_crazaodestinatario', 'cabecalho_ncodnf']
    list_per_page = 20
    readonly_fields = ['sync_created_at', 'sync_updated_at']


# Configuração para PedidoVendaItens
@admin.register(PedidoVendaItens)
class PedidoVendaItensAdmin(admin.ModelAdmin):
    list_display = ['id', 'parent_id', 'produto_codigo_produto', 'produto_descricao', 'produto_quantidade', 'produto_valor_total']
    list_filter = ['produto_cfop', 'produto_reservado']
    search_fields = ['produto_descricao', 'produto_codigo']
    list_per_page = 20
    readonly_fields = ['sync_created_at', 'sync_updated_at']


# Configuração para PedidoVendaProduto
@admin.register(PedidoVendaProduto)
class PedidoVendaProdutoAdmin(admin.ModelAdmin):
    list_display = ['cabecalho_codigo_pedido', 'cabecalho_numero_pedido', 'data_emissao', 'cliente_fantasia', 'cliente_razao_social', 'cliente_cnpj', 'valor_sem_frete', 'nome_projeto', 'nome_vendedor', 'status_pedido']
    list_filter = [
        'cabecalho_encerrado', 
        'cabecalho_etapa', 
        'cabecalho_origem_pedido', 
        'cabecalho_bloqueado', 
        'infocadastro_faturado', 
        'infocadastro_cancelado',
        ('cabecalho_data_previsao', admin.DateFieldListFilter)
    ]
    search_fields = ['cabecalho_numero_pedido', 'cabecalho_codigo_pedido_integracao', 'cliente__razao_social', 'vendedor__nome']
    list_per_page = 25
    ordering = ['-cabecalho_numero_pedido']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    autocomplete_fields = ['cliente', 'vendedor', 'projeto']
    list_select_related = ['cliente', 'vendedor', 'projeto']
    actions = [export_to_excel]
    
    @admin.display(description='Data Emissão')
    def data_emissao(self, obj):
        return obj.infocadastro_dinc or '-'
    
    @admin.display(description='Cliente Fantasia')
    def cliente_fantasia(self, obj):
        if obj.cliente:
            return obj.cliente.nome_fantasia or '-'
        return '-'
    
    @admin.display(description='Razão Social')
    def cliente_razao_social(self, obj):
        if obj.cliente:
            return obj.cliente.razao_social or '-'
        return '-'
    
    @admin.display(description='CNPJ/CPF')
    def cliente_cnpj(self, obj):
        if obj.cliente:
            return obj.cliente.cnpj_cpf or '-'
        return '-'
    
    @admin.display(description='Valor s/ Frete')
    def valor_sem_frete(self, obj):
        if obj.total_pedido_valor_mercadorias:
            return f'R$ {obj.total_pedido_valor_mercadorias:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
        return '-'
    
    @admin.display(description='Vendedor')
    def nome_vendedor(self, obj):
        if obj.vendedor:
            return obj.vendedor.nome
        return '-'
    
    @admin.display(description='Projeto')
    def nome_projeto(self, obj):
        if obj.projeto:
            return obj.projeto.nome
        return '-'
    
    @admin.display(description='Status')
    def status_pedido(self, obj):
        if obj.infocadastro_cancelado == 'S':
            return mark_safe('<span style="color: red;">● Cancelado</span>')
        if obj.infocadastro_faturado == 'S':
            return mark_safe('<span style="color: green;">● Faturado</span>')
        if obj.cabecalho_encerrado == 'S':
            return mark_safe('<span style="color: blue;">● Encerrado</span>')
        if obj.cabecalho_bloqueado == 'S':
            return mark_safe('<span style="color: orange;">● Bloqueado</span>')
        return mark_safe('<span style="color: gray;">● Em Aberto</span>')
    
    fieldsets = (
        ('Cabeçalho do Pedido', {
            'fields': ('cabecalho_codigo_pedido', 'cabecalho_numero_pedido', 'cliente', 'cabecalho_data_previsao')
        }),
        ('Responsáveis', {
            'fields': ('vendedor', 'projeto')
        }),
        ('Status', {
            'fields': ('cabecalho_encerrado', 'cabecalho_etapa', 'cabecalho_bloqueado')
        }),
        ('Valores', {
            'fields': ('total_pedido_valor_mercadorias', 'total_pedido_valor_total_pedido', 'total_pedido_valor_descontos')
        }),
    )


# Configuração para ProjetosCadastro
@admin.register(ProjetosCadastro)
class ProjetosCadastroAdmin(admin.ModelAdmin):
    list_display = ['codigo', 'nome', 'codint', 'status_projeto']
    list_filter = ['inativo']
    search_fields = ['nome', 'codint']
    list_per_page = 25
    ordering = ['nome']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    fieldsets = (
        ('📂 Dados do Projeto', {
            'fields': (
                ('codigo', 'codint'),
                ('nome', 'inativo'),
            ),
            'classes': ('wide',),
            'description': 'Informações principais do projeto'
        }),
        ('🕰️ Sistema', {
            'fields': (
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )
    
    @admin.display(description='Status')
    def status_projeto(self, obj):
        if obj.inativo == 'S':
            return mark_safe('<span style="color: red;">● Inativo</span>')
        return mark_safe('<span style="color: green;">● Ativo</span>')


# Configuração para VendedoresCadastro
@admin.register(VendedoresCadastro)
class VendedoresCadastroAdmin(admin.ModelAdmin):
    list_display = ['codigo', 'nome', 'email', 'comissao_formatada', 'status_vendedor']
    list_filter = [
        'inativo', 
        'fatura_pedido', 
        'visualiza_pedido',
        ('email', admin.EmptyFieldListFilter),
        ('comissao', admin.EmptyFieldListFilter)
    ]
    search_fields = ['nome', 'email', 'codint']
    list_per_page = 25
    ordering = ['nome']
    readonly_fields = ['sync_created_at', 'sync_updated_at']
    
    @admin.display(description='Comissão')
    def comissao_formatada(self, obj):
        if obj.comissao:
            return f'{obj.comissao}%'
        return '-'
    
    @admin.display(description='Status')
    def status_vendedor(self, obj):
        if obj.inativo == 'S':
            return mark_safe('<span style="color: red;">● Inativo</span>')
        return mark_safe('<span style="color: green;">● Ativo</span>')
    
    fieldsets = (
        ('👥 Dados do Vendedor', {
            'fields': (
                ('codigo', 'codint'),
                ('nome', 'email'),
                ('inativo',),
            ),
            'classes': ('wide',),
            'description': 'Informações básicas do vendedor'
        }),
        ('💼 Config. Comerciais', {
            'fields': (
                ('comissao',),
                ('fatura_pedido', 'visualiza_pedido'),
            ),
            'classes': ('wide',)
        }),
        ('🕰️ Sistema', {
            'fields': (
                ('sync_created_at', 'sync_updated_at'),
            ),
            'classes': ('wide', 'collapse')
        }),
    )


# Configuração do Admin Site
admin.site.site_header = "OMIE - Administração"
admin.site.site_title = "OMIE Admin"
admin.site.index_title = "Sistema de Administração OMIE"
